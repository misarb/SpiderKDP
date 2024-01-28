from PyQt5.QtGui import *
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
import sys
from PyQt5.uic import loadUiType
import tkinter as tk
from tkinter import filedialog
import pickle as pk
import os.path
from os import path
import os
import glob
import time
import random
import json
from fake_useragent import UserAgent
from random import randint
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import webbrowser
import openpyxl
import pandas as pd
from xlwt import Workbook


########################## Variables ##########################################
path_xl = 0
kdp_login_link1="https://kdp.amazon.com/en_US/title-setup/paperback/new/details?openid.assoc_handle=amzn_dtp&openid.claimed_id=https%3A%2F%2Fwww.amazon.com%2Fap%2Fid%2Famzn1.account.AG7HI6NMSZIPA7QVZPQKJ3I6XJOQ&openid.identity=https%3A%2F%2Fwww.amazon.com%2Fap%2Fid%2Famzn1.account.AG7HI6NMSZIPA7QVZPQKJ3I6XJOQ&openid.mode=id_res&openid.ns=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0&openid.op_endpoint=https%3A%2F%2Fwww.amazon.com%2Fap%2Fsignin&openid.response_nonce=2021-01-01T14%3A40%3A13Z5054585209589897938&openid.return_to=https%3A%2F%2Fkdp.amazon.com%2Fen_US%2Ftitle-setup%2Fpaperback%2Fnew%2Fdetails&openid.signed=assoc_handle%2Cclaimed_id%2Cidentity%2Cmode%2Cns%2Cop_endpoint%2Cresponse_nonce%2Creturn_to%2CsiteState%2Cns.pape%2Cpape.auth_policies%2Cpape.auth_time%2Csigned&openid.ns.pape=http%3A%2F%2Fspecs.openid.net%2Fextensions%2Fpape%2F1.0&openid.pape.auth_policies=http%3A%2F%2Fschemas.openid.net%2Fpape%2Fpolicies%2F2007%2F06%2Fnone&openid.pape.auth_time=2021-01-01T14%3A40%3A13Z&openid.sig=VF%2FUmQ8Gn%2BJZlFN2SKJyfiLz4tapfSLlYsIaam4nMw8%3D&serial=&siteState=clientContext%3D146-2350627-9370254%2CsourceUrl%3Dhttps%253A%252F%252Fkdp.amazon.com%252Fen_US%252Ftitle-setup%252Fpaperback%252Fnew%252Fdetails%253Fref_%253Dkdp_kdp_BS_D_cr_ti%2526ref_%253Dkdp_kdp_BS_D_cr_ti%2Csignature%3DM0gmTWdi1NsM4Ej2BzYD6zBKd5oFgj3D"
kdp_login_link = "https://www.amazon.com/ap/signin?clientContext=130-9897631-2090763&openid.return_to=https%3A%2F%2Fkdp.amazon.com%2Fap-post-redirect&openid.identity=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0%2Fidentifier_select&openid.assoc_handle=amzn_dtp&openid.mode=checkid_setup&marketPlaceId=ATVPDKIKX0DER&openid.claimed_id=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0%2Fidentifier_select&pageId=kdp-ap&openid.ns=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0&language=en_US&openid.pape.max_auth_age=0&siteState=clientContext%3D140-4224367-4041222%2CsourceUrl%3Dhttps%253A%252F%252Fkdp.amazon.com%252Fen_US%252Fbookshelf%2Csignature%3DxeKDgkLLOYA73oDBjvam2SlRpXoj3D"
kdp_signin_button = "//input[@id='signInSubmit']"
kdp_username_input = "//input[@id='ap_email']"
kdp_pass_input = "//input[@id='ap_password']"
kdp_login_button = "//input[@id='signInSubmit']"
kdp_test = "//body/div[@id='a-page']/div[@id='page-container']/div[5]/div[1]/div[2]/div[1]/div[5]/div[2]/div[1]/div[1]/div[1]/a[2]/span[1]/span[1]/input[1]"
kdp_search="//span[@id='podbookshelftable-search-button-submit-announce']"
kdp_test_1="//input[@id='data-print-book-title']"
kdp_subtitle="//input[@id='data-print-book-subtitle']"
kdp_prefix="//input[@id='data-print-book-primary-author-prefix']"
kdp_firstname="//input[@id='data-print-book-primary-author-first-name']"
kdp_Middle_name="//input[@id='data-print-book-primary-author-middle-name']"
kdp_last_name="//input[@id='data-print-book-primary-author-last-name']"
kdp_suffix="//input[@id='data-print-book-primary-author-suffix']"
kdp_Description="//body[@class='cke_editable cke_editable_themed cke_contents_ltr cke_show_borders']"
kdp_keyword_1="//input[@id='data-print-book-keywords-0']"
kdp_keyword_2="//input[@id='data-print-book-keywords-2']"
kdp_keyword_3="//input[@id='data-print-book-keywords-4']"
kdp_keyword_4="//input[@id='data-print-book-keywords-6']"
kdp_keyword_5="//input[@id='data-print-book-keywords-5']"
kdp_keyword_6="//input[@id='data-print-book-keywords-3']"
kdp_keyword_7="//input[@id='data-print-book-keywords-1']"
kdp_adult_content="//body/div[@id='a-page']/div[@id='page-container']/div[3]/div[2]/div[2]/form[1]/div[1]/div[13]/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/div[1]/span[1]/span[1]/span[1]/div[1]/div[1]/div[1]/div[1]/div[1]/label[1]/input[1]"
kdp_publish_right ="//input[@id='public-domain']"
kdp_save_category="//span[@id='category-chooser-ok-button']/span/input"
kdp_save_page1="//button[@id='save-and-continue-announce']"
######################### Variable for page 2 ####################################
kdp_isbn="//button[@id='free-print-isbn-btn-announce']"
kdp_asign_isbn="//button[@id='print-isbn-confirm-button-announce']"
kdp_succses_isbn="//div[@id='print-isbn-success-alert']/div/i"
kdp_paperback_cream="//button[@id='a-autoid-0-announce']"
kdp_paperback_white="//button[@id='a-autoid-1-announce']/span"
kdp_no_bleed="//button[@id='a-autoid-4-announce']"
kdp_bleed="//button[@id='a-autoid-5-announce']"
kdp_paperback_Matte="//button[@id='a-autoid-6-announce']"
kdp_paperback_Glossy="//button[@id='a-autoid-7-announce']"
kdp_upload_Button_manuscript="//button[@id='data-print-book-publisher-interior-file-upload-browse-button-announce']"
kdp_manuscript_succses="//div[@id='data-print-book-publisher-interior-file-upload-success']/div/i"
kdp_upload_path="id=data-print-book-publisher-interior-file-upload-AjaxInput"
kdp_upload_cover_choice="//div[@id='data-print-book-publisher-cover-choice-accordion']/div[2]/div/div/a/h5"
kdp_uploadcover_button="//button[@id='data-print-book-publisher-cover-file-upload-browse-button-announce']"
kdp_upload_cover_path="//input[@id='data-print-book-publisher-cover-file-upload-AjaxInput']"
kdp_cover_succses="//div[@id='data-print-book-publisher-cover-file-upload-success']/div/i"
kdp_cover_barcode="//input[@id='data-print-book-has-publisher-barcode']"
kdp_lunch_preview="//button[@id='print-preview-noconfirm-announce']"
kdp_approve_preview="//a[contains(text(),'Approve')]"
kdp_save_two="//button[@id='save-and-continue-announce']"
############################# Page 3 #####################
kdp_price="//input[@name='data[print_book][amazon_channel][us][price_vat_exclusive]']"
kdp_price_world="//*[@id='print-proofs-copies']/div/i"
kdp_save_draft="//button[@id='save-announce']"
############################### variable excel ##############################""
column_book_title=3
column_book_subtitle=4
column_auther=5
column_description=10
column_keyword=11
column_interior_path=1
column_cover_path=2
column_price=18



ui, _ = loadUiType('spiderkdp.ui')


class MainApp(QMainWindow, ui):

    def __init__(self, parent=None):
        super(MainApp, self).__init__(parent)
        QMainWindow.__init__(self)
        self.setupUi(self)
        self.Handel_Buttons()
        self.load_data_init()
        

       

    def handle_browser(self):
        global path_xl
        excel_file = QFileDialog.getOpenFileName(self, 'Open file', 'c:\\', "Excel files (*.xlsx)")
        self.excel_path.setText(excel_file[0])
        path_xl = excel_file[0]


    ################# Generate Excel file with Header Forma  ######################
    def GenerateExcelFormate(self):
     data = []
        #  Tables for Headears Data
     df = pd.DataFrame({'Interior File Path': data,
         'Cover File Path' : data,
         'Title' : data,
         'Subtitle' : data,
         'Prefix' : data,
         'First Name' : data,
         'Middle Name' : data,
         'Last Name' : data,
         'Suffix' : data,
         'Description':data,
         'Keyword #1' : data,
         'Keyword #2' : data,
         'Keyword #3' : data,
         'Keyword #4' : data,
         'Keyword #5' : data,
         'Keyword #6' : data,
         'Keyword #7' : data,
         'Price' : data,
         'Category #1 (Optional)' : data,
         'Category #2 (Optional)' : data,
                   })

      
      #  Excel writer using XlsxWriter 
     #writer = pd.ExcelWriter("SPIDER_KDP_V1_format.xlsx", engine='xlsxwriter')
     desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), r'Desktop\Forma_SpiderKdp_V1.xlsx')

     #Header
     print(desktop)
     writer=pd.ExcelWriter(desktop,engine='xlsxwriter')
     df.to_excel(writer, sheet_name='Sheet1', startrow=0, header=False) 
     

     #  worksheet objects.
     workbook  = writer.book
     worksheet = writer.sheets['Sheet1']

     #  header format.
     header_format = workbook.add_format({
        'bold': False,
        'text_wrap': True,
        'valign': 'top',
        'fg_color': '#D7E4BC',
        'border': 1
        })

     #  headers with the defined format.
     for col_num, value in enumerate(df.columns.values):
         worksheet.write(0, col_num , value, header_format)
    

     #Save Data 
     writer.save()

     QMessageBox.information(self, "SpiderKDP 1.0", " The Excel fille is Generated ")
  


       

        

    def taps_loads(self):
        
         if (self.excel_path.text() == ""):
             QMessageBox.information(self, "SpiderKDP 1.0", " Choose an Excel file first ")
             return

         EmptyCheck=0
         isokay = True
         nbr_tabs = self.number_tabs.value()
         xl = openpyxl.load_workbook(path_xl)
         sheet_obj = xl.active
         max_row = sheet_obj.max_row
         max_upload = self.max_upload.value()
         if max_upload >=  max_row or max_row < nbr_tabs:
             isokay=False
             QMessageBox.information(self, "SpiderKDP 1.0", "There's something wrong with excel file !!, you need to check if numbers of liness in excel is larger than max upload and number of tabs or the excel fille is empty")               
         elif max_upload <=  max_row and max_row > nbr_tabs:

              ######## check empty line in excel  ##########
             for i in range(1,19):
                 for j in range(1,max_row):
                     if sheet_obj.cell(row=j,column=i).value == None:
                         isokay=True
                         EmptyCheck=EmptyCheck + 1
                          
        #  if EmptyCheck > 0:
        #      QMessageBox.information(self, "SpiderKDP 1.0", "some lines in excel fille is empty please fill the void") 
                            
         if isokay==True:
            QMessageBox.information(self, "SpiderKDP 1.0", "You can LAUNCH NOW!") #### move to the next button 
            self.launch_button.setEnabled(True)

    def Login(self): 
        shut_down = self.radio_yes.isChecked() and not self.radio_no.isChecked()
        

        ####################### Agent ######################
        chrome_options = Options()
        # ua = UserAgent()
        # userAgent=ua.random
        # chrome_options.add_argument(f'user-agent={userAgent}')
        chrome_options.add_argument("user-data-dir=C:\\Users\\python\\AppData\\Local\\Google\\Chrome Beta\\User Data")
        chrome_options.binary_location = "C:\\Program Files\\Google\\Chrome Beta\\Application\\chrome.exe"
        driver = webdriver.Chrome(executable_path=ChromeDriverManager().install(),options=chrome_options)
        nbr_tabs = self.number_tabs.value()
        # except Exception as ex:
        #     QMessageBox.information(self, "SpiderKDP 1.0", "" +str(ex)) 
        #     sys.exit()
            


        
        ###  Opening tabs in browser #####
        for i in range(nbr_tabs):
            driver.execute_script("window.open('about:blank','tab"+str(i)+"'"+");")

        #####  filed and switcch between tabs  #########

        for i in range(nbr_tabs):
            driver.switch_to.window("tab"+str(i))
            self.AutomatedPags(driver,i)
            time.sleep(2)

        ##### shutDown Computer   #####
        
        if shut_down==True:
            os.system("shutdown /s /t 1")    
        else:
            sys.exit()

  
     


    def AutomatedPags(self,web,i):
        driver=web
        try: 
        #   if i==0:
        #       wait=WebDriverWait(driver,20000)
        #       driver.get(kdp_login_link)
        #       driver.maximize_window()
        #       wait.until(EC.element_to_be_clickable((By.XPATH,kdp_test))).click()
        #   else:
            driver.get(kdp_login_link1)
        except Exception as ex:
            QMessageBox.information(self, "SpiderKDP 1.0", "" +str(ex)) 
            sys.exit()


        

      
        ##########************** Start Code to Automate the pages ***********************###########
        
        wait=WebDriverWait(driver,20000)
        wait.until(EC.element_to_be_clickable((By.XPATH,kdp_test_1))).click()
        test_1=driver.find_element_by_xpath(kdp_test_1)
        xl = openpyxl.load_workbook(path_xl)
        sheet_obj = xl.active
        test_1.send_keys(sheet_obj.cell(row=i+2,column=column_book_title).value )
        subtitle=driver.find_element_by_xpath(kdp_subtitle)
        subtitle.send_keys(sheet_obj.cell(row=i+2,column=column_book_subtitle).value)
        time.sleep(3)
        # prefix=driver.find_element_by_xpath(kdp_prefix)
        # prefix.send_keys(sheet_obj.cell(row=i+2,column=column_auther).value)
        first_name=driver.find_element_by_xpath(kdp_firstname)
        first_name.send_keys(sheet_obj.cell(row=i+2,column=column_auther+1).value)
        # middle_name=driver.find_element_by_xpath(kdp_Middle_name)
        # middle_name.send_keys(sheet_obj.cell(row=i+2,column=column_auther+2).value)
        last_name=driver.find_element_by_xpath(kdp_last_name)
        last_name.send_keys(sheet_obj.cell(row=i+2,column=column_auther+3).value)
        time.sleep(3)
        # suffix=driver.find_element_by_xpath(kdp_suffix)
        # suffix.send_keys(sheet_obj.cell(row=i+2,column=column_auther+4).value)
        iframe = driver.find_element_by_xpath("//iframe[@class='cke_wysiwyg_frame cke_reset']")
        driver.switch_to.frame(iframe)
        description=driver.find_element_by_xpath(kdp_Description)
        description.click()
        description.send_keys(sheet_obj.cell(row=i+2,column=column_description).value)
        driver.switch_to.default_content()
        time.sleep(2)
        publishing_right=driver.find_element_by_xpath(kdp_publish_right)
        publishing_right.click()
        time.sleep(2)
        keyword_1=driver.find_element_by_xpath(kdp_keyword_1)
        keyword_2=driver.find_element_by_xpath(kdp_keyword_2)
        keyword_3=driver.find_element_by_xpath(kdp_keyword_3)
        keyword_4=driver.find_element_by_xpath(kdp_keyword_4)
        keyword_5=driver.find_element_by_xpath(kdp_keyword_5)
        keyword_6=driver.find_element_by_xpath(kdp_keyword_6)
        keyword_7=driver.find_element_by_xpath(kdp_keyword_7)
        keyword_1.send_keys(sheet_obj.cell(row=i+2,column=column_keyword).value)
        keyword_2.send_keys(sheet_obj.cell(row=i+2,column=column_keyword+1).value)
        keyword_3.send_keys(sheet_obj.cell(row=i+2,column=column_keyword+2).value)
        keyword_4.send_keys(sheet_obj.cell(row=i+2,column=column_keyword+3).value)
        keyword_5.send_keys(sheet_obj.cell(row=i+2,column=column_keyword+4).value)
        keyword_6.send_keys(sheet_obj.cell(row=i+2,column=column_keyword+5).value)
        keyword_7.send_keys(sheet_obj.cell(row=i+2,column=column_keyword+6).value)
        time.sleep(2)
        ######################## Categories ############################
        category_list=sheet_obj.cell(row=i+2,column=19).value
        # scro_cat=driver.find_element_by_xpath("//button[contains(text(),'Choose categories')]")
        # scro_cat.location_once_scrolled_into_view

        X=category_list.split(" > ")
        wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(),'Choose categories')]"))).click()
        time.sleep(2)
        wait.until(EC.element_to_be_clickable((By.XPATH, "//a[text()="+"'"+ X[0] + "'"+"]"))).click()
        wait.until(EC.element_to_be_clickable((By.XPATH, "//span[text()=" +"'"    +X[1]+  "'"+    "]//preceding-sibling::input"))).click()
        time.sleep(2)
        wait.until(EC.element_to_be_clickable((By.XPATH, kdp_save_category))).click()
        time.sleep(2)
        driver.find_element_by_xpath("//*[@id='data-view-is-lcb']").click()
        time.sleep(2)
        ############################## End Categories ###################################
        Adult_Ans=self.adult_content.currentText()
        if Adult_Ans=="No":
             wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@name='data[print_book][is_adult_content]-radio']"))).click()
        else:
             wait.until(EC.element_to_be_clickable((By.XPATH, "(//input[@name='data[print_book][is_adult_content]-radio'])[2]"))).click()

        time.sleep(2)
        wait.until(EC.element_to_be_clickable((By.ID,"save-and-continue-announce"))).click()
        ############################## End Page 1 ###################################
        # wait.until(EC.element_to_be_clickable((By.XPATH, kdp_isbn))).click()
        # time.sleep(3)
        # wait.until(EC.element_to_be_clickable((By.XPATH, kdp_asign_isbn))).click()
        # wait.until(EC.element_to_be_clickable((By.XPATH, kdp_succses_isbn))).click()
        paperback_cover=self.paper_type.currentText()
        if paperback_cover=='Cream' :
            wait.until(EC.element_to_be_clickable((By.XPATH, kdp_paperback_cream))).click()
        else:
            wait.until(EC.element_to_be_clickable((By.XPATH, kdp_paperback_white))).click()

        AnsBleed=self.bleed.currentText()
        if AnsBleed=='Bleed':
            wait.until(EC.element_to_be_clickable((By.XPATH, kdp_bleed))).click()
        else:
            wait.until(EC.element_to_be_clickable((By.XPATH, kdp_no_bleed))).click()

        finish_cover=self.cover_finish.currentText()
        if finish_cover=="Matte":
            wait.until(EC.element_to_be_clickable((By.XPATH, kdp_paperback_Matte))).click()
        else:
            wait.until(EC.element_to_be_clickable((By.XPATH, kdp_paperback_Glossy))).click()
        
        driver.find_element_by_id("data-print-book-publisher-interior-file-upload-AjaxInput").send_keys(sheet_obj.cell(row=i+2,column=column_interior_path).value)
        wait.until(EC.element_to_be_clickable((By.XPATH, kdp_manuscript_succses))).click()
        wait.until(EC.element_to_be_clickable((By.XPATH, kdp_upload_cover_choice))).click()
        driver.find_element_by_xpath(kdp_upload_cover_path).send_keys(sheet_obj.cell(row=i+2,column=column_cover_path).value)
        wait.until(EC.element_to_be_clickable((By.XPATH, kdp_cover_succses))).click()
        # scro=driver.find_element_by_xpath(kdp_upload_cover_choice)
        # scro.location_once_scrolled_into_view
        # wait.until(EC.element_to_be_clickable((By.XPATH, kdp_cover_barcode))).click()
        wait.until(EC.element_to_be_clickable((By.XPATH, kdp_lunch_preview))).click()
        wait.until(EC.element_to_be_clickable((By.XPATH, kdp_approve_preview))).click()
        wait.until(EC.element_to_be_clickable((By.XPATH, kdp_save_two))).click()
        wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@id='data-print-book-worldwide-rights-accordion']/div/div/div/a/i"))).click()
        driver.find_element_by_xpath(kdp_price).send_keys(sheet_obj.cell(row=2,column=column_price).value)
        wait.until(EC.element_to_be_clickable((By.XPATH, kdp_price_world))).click()
        AnsPublsh=self.publish.currentText()
        if AnsPublsh=='No':
            wait.until(EC.element_to_be_clickable((By.XPATH, kdp_save_draft))).click()
        else:
            wait.until(EC.element_to_be_clickable((By.XPATH, kdp_save_draft))).click()

             

        

  
      

    def load_data_init(self):
        self.bleed.addItems(["Bleed", "No Bleed"])
        self.publish.addItems(["Publish", "Draft"])
        self.paper_type.addItems(["Cream", "White"])
        self.cover_finish.addItems(["Matte", "Glossy"])
        self.adult_content.addItems(["No", "Yes"])


    def go_to_facebookgroup(self):
        webbrowser.open('https://www.facebook.com/groups/345254530051292/', new=2)
        
 

    def Handel_Buttons(self):
        self.facebook_button.clicked.connect(self.go_to_facebookgroup)
        self.browse_button.clicked.connect(self.handle_browser)
        self.launch_button.clicked.connect(self.Login)
        self.export_butn.clicked.connect(self.GenerateExcelFormate)
        self.check_button.clicked.connect(self. taps_loads)
        
        

def main():
    app = QApplication(sys.argv)
    window = MainApp()
    window.show()
    app.exec_()


if __name__ == '__main__':
    main()



